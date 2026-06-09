"""Local S2 scoring based on Excel definitions, event CSVs, and market data."""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field, replace
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from s2.event_store import ensure_event_store, load_events
from s2.market_metrics import ClinicalEventStatus, MarketResult, clinical_conversion_rate, leader_excess_median


DEFAULT_DEFINITIONS = {
    "S2-01": ("BD落地频率", 0.2, ">=1.5x", "0.8x~1.5x", "<0.8x"),
    "S2-02": ("BD金额质量", 0.2, ">=150%", "80%~150%", "<80%"),
    "S2-03a": ("财报客观改善", 0.1, ">=70%", "40%~70%", "<40%"),
    "S2-03b": ("一致预期验证", 0.1, ">=60%", "40%~60%", "<40%"),
    "S2-04": ("数据催化转化率", 0.2, ">=60%", "40%~60%", "<40%"),
    "S2-05": ("龙头接力强度", 0.2, ">=5%", "0%~5%", "<0%"),
    "S2-06": ("商业化兑现质量", 0.0, ">=70%", "40%~70%", "<40%"),
}


@dataclass(frozen=True)
class S2Definition:
    code: str
    name: str
    weight: float
    exceed: str
    meet: str
    below: str


@dataclass(frozen=True)
class S2Item:
    code: str
    name: str
    value: float | None
    raw_score: float
    adjusted_score: float
    confidence: float
    rating: str
    basis: str
    source: str
    sample_count: int = 0
    replacement_count: int = 0
    missing: str = ""
    event_db_maturity: str = ""
    raw_bd_amount: float | None = None
    quality_bd_amount: float | None = None
    baseline_bd_amount: float | None = None
    true_value: float | None = None
    proxy_value: float | None = None
    true_sample_count: int = 0
    proxy_sample_count: int = 0
    proxy_type: str = ""
    leader_excess_median_5d: float | None = None
    leader_win_rate_5d: float | None = None
    leader_excess_median_10d: float | None = None
    leader_breadth_20d: float | None = None
    pending_count: int = 0
    hk_pending_count: int = 0
    price_missing_count: int = 0
    raw_mature_event_count: int = 0
    deduped_trade_sample_count: int = 0
    success_count: int = 0
    success_rate: float | None = None
    carried_forward_from: str = ""
    stale_days: int = 0
    is_stale: bool = False
    carry_forward_type: str = ""
    clinical_event_statuses: tuple[ClinicalEventStatus, ...] = ()

    @property
    def score(self) -> float:
        """Backward-compatible official score alias."""
        return self.adjusted_score


@dataclass(frozen=True)
class S2Score:
    trade_date: str
    raw_score: float
    adjusted_score: float
    items: dict[str, S2Item]
    missing_data: str
    level: str
    available_weight: float
    missing_indicator_count: int
    pending_indicator_count: int
    proxy_indicator_count: int
    stale_indicator_count: int
    explanation_items: dict[str, S2Item] = field(default_factory=dict)

    @property
    def total_score(self) -> float:
        """Backward-compatible official total score alias."""
        return self.adjusted_score


def load_definitions(excel_path: Path) -> dict[str, S2Definition]:
    if not excel_path.exists():
        return {
            code: S2Definition(code, name, weight, exceed, meet, below)
            for code, (name, weight, exceed, meet, below) in DEFAULT_DEFINITIONS.items()
        }
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb["第二阶段"]
    definitions: dict[str, S2Definition] = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        code = row[1]
        if not code or not str(code).startswith("S2-"):
            continue
        definitions[str(code)] = S2Definition(
            code=str(code),
            name=str(row[3]),
            weight=float(row[11] or 0.2),
            exceed=str(row[8]),
            meet=str(row[9]),
            below=str(row[10]),
        )
    if not definitions:
        return {
            code: S2Definition(code, name, weight, exceed, meet, below)
            for code, (name, weight, exceed, meet, below) in DEFAULT_DEFINITIONS.items()
        }
    for code, (name, weight, exceed, meet, below) in DEFAULT_DEFINITIONS.items():
        definitions.setdefault(code, S2Definition(code, name, weight, exceed, meet, below))
    return definitions


def _default_definitions() -> dict[str, S2Definition]:
    return {
        code: S2Definition(code, name, weight, exceed, meet, below)
        for code, (name, weight, exceed, meet, below) in DEFAULT_DEFINITIONS.items()
    }


def _parse_date(value: str) -> datetime | None:
    raw = value.replace("-", "")
    try:
        return datetime.strptime(raw, "%Y%m%d")
    except ValueError:
        return None


def _is_confirmed_active(event: dict[str, str]) -> bool:
    return (
        event.get("status", "active") == "active"
        and event.get("verification_status", "confirmed") == "confirmed"
        and event.get("is_duplicate", "false").lower() not in {"true", "1", "yes", "y", "是"}
    )


def _active_since(events: list[dict[str, str]], trade_date: str, days: int) -> list[dict[str, str]]:
    end = _parse_date(trade_date)
    if end is None:
        return []
    start = end - timedelta(days=days)
    active = []
    for event in events:
        event_date = _parse_date(event.get("date", ""))
        if event_date and start <= event_date <= end and _is_confirmed_active(event):
            active.append(event)
    return active


def _active_between(
    events: list[dict[str, str]],
    trade_date: str,
    start_days_ago: int,
    end_days_ago: int,
) -> list[dict[str, str]]:
    end = _parse_date(trade_date)
    if end is None:
        return []
    start = end - timedelta(days=start_days_ago)
    stop = end - timedelta(days=end_days_ago)
    return [
        event
        for event in events
        if (event_date := _parse_date(event.get("date", "")))
        and start <= event_date < stop
        and _is_confirmed_active(event)
    ]


def _major_bd(events: list[dict[str, str]]) -> list[dict[str, str]]:
    return [event for event in events if event.get("is_major_bd", "true").lower() in {"true", "1", "yes", "y", "是"}]


def _market_events_since(
    events: list[dict[str, str]],
    trade_date: str,
    days: int,
    report_date: str | None = None,
) -> list[dict[str, str]]:
    end = _parse_date(trade_date)
    if end is None:
        return []
    cutoff = _parse_date(report_date or trade_date) or end
    start = end - timedelta(days=days)
    return [
        event
        for event in events
        if (event_date := _parse_date(event.get("effective_trade_date") or event.get("date", "")))
        and start <= event_date <= cutoff
        and event.get("status", "active") in {"active", "superseded"}
        and event.get("verification_status", "confirmed") == "confirmed"
        and event.get("is_duplicate", "false").lower() not in {"true", "1", "yes", "y", "是"}
    ]


def _number(value: str) -> float:
    try:
        return float(value or 0)
    except ValueError:
        return 0.0


def _rate_score(value: float | None, exceed: float, meet: float) -> tuple[float, str]:
    if value is None:
        return 0.5, "数据缺失"
    if value >= exceed:
        return 1.0, "超预期"
    if value >= meet:
        return 0.7, "符合预期"
    return 0.4, "低于预期"


def _confidence(sample_count: int, replacement_count: int, missing: str = "") -> float:
    if sample_count <= 0:
        base = 0.45
    elif sample_count < 3:
        base = 0.60
    elif sample_count < 5:
        base = 0.75
    else:
        base = 0.90
    if replacement_count > 0:
        base = min(base, 0.65)
    if missing:
        base = min(base, 0.70)
    return base


def _adjust(raw_score: float, sample_count: int, replacement_count: int = 0, missing: str = "") -> tuple[float, float]:
    confidence = _confidence(sample_count, replacement_count, missing)
    cap = 1.0
    if sample_count < 3:
        cap = min(cap, 0.65)
    elif sample_count < 5:
        cap = min(cap, 0.75)
    if replacement_count > 0:
        cap = min(cap, 0.70)
    if missing and sample_count <= 0:
        cap = min(cap, 0.60)
    return min(raw_score, cap), confidence


def _event_db_maturity(count_365d: int) -> str:
    if count_365d < 8:
        return "low"
    if count_365d < 15:
        return "medium"
    return "high"


def _bd_frequency(defn: S2Definition, bd_events: list[dict[str, str]], trade_date: str) -> S2Item:
    recent = _major_bd(_active_since(bd_events, trade_date, 90))
    baseline_windows = [
        _major_bd(_active_between(bd_events, trade_date, start, stop))
        for start, stop in [(180, 90), (270, 180), (360, 270), (450, 360)]
    ]
    history = _major_bd(_active_since(bd_events, trade_date, 365))
    baseline = sum(len(window) for window in baseline_windows) / 4
    if baseline <= 0:
        value = None
        raw_score, rating = 0.5, "数据缺失"
        missing = "前4个完整90日窗口无重大BD基准，S2-01不评分"
    else:
        value = len(recent) / baseline
        raw_score, rating = _rate_score(value, 1.5, 0.8)
        missing = ""
    baseline_counts = "/".join(str(len(window)) for window in baseline_windows)
    basis = f"近90日重大BD {len(recent)} 笔；前4个完整90日窗口 {baseline_counts} 笔，单窗口均值 {baseline:.2f}"
    adjusted, confidence = _adjust(raw_score, len(recent), missing=missing)
    maturity = _event_db_maturity(len(history))
    if maturity == "low":
        adjusted = min(adjusted, 0.70)
    elif maturity == "medium":
        adjusted = min(adjusted, 0.80)
    return S2Item(defn.code, defn.name, value, raw_score, adjusted, confidence, rating, basis, "bd_events.csv", len(recent), 0, missing, maturity)


def _bd_quality(defn: S2Definition, bd_events: list[dict[str, str]], trade_date: str) -> S2Item:
    recent = _major_bd(_active_since(bd_events, trade_date, 90))
    baseline_events = _major_bd(_active_between(bd_events, trade_date, 455, 365))
    amount = sum(_number(e.get("upfront_usd", "")) + _number(e.get("near_term_milestone_usd", "")) for e in recent)
    quality_amount = sum(
        _number(e.get("upfront_usd", ""))
        + 0.5 * _number(e.get("near_term_milestone_usd", ""))
        + 0.1 * _number(e.get("long_term_milestone_usd", ""))
        for e in recent
    )
    baseline_amount = sum(_number(e.get("upfront_usd", "")) + _number(e.get("near_term_milestone_usd", "")) for e in baseline_events)
    if baseline_amount <= 0:
        value = None
        raw_score, rating = 0.5, "数据缺失"
        missing = "去年同期90日BD金额基准缺失，S2-02不评分"
    else:
        value = amount / baseline_amount
        raw_score, rating = _rate_score(value, 1.50, 0.80)
        missing = ""
    adjusted, confidence = _adjust(raw_score, len(recent), missing=missing)
    basis = f"近90日raw金额 {amount:,.0f} USD；去年同期90日 {baseline_amount:,.0f} USD；质量金额 {quality_amount:,.0f} USD"
    return S2Item(defn.code, defn.name, value, raw_score, adjusted, confidence, rating, basis, "bd_events.csv", len(recent), 0, missing, raw_bd_amount=amount, quality_bd_amount=quality_amount, baseline_bd_amount=baseline_amount)


def _boolish(value: str) -> bool:
    return value.lower() in {"true", "1", "yes", "y", "是"}


def _positive_number(value: str) -> bool:
    raw = str(value or "").strip().lower()
    if raw in {"turnaround", "loss_narrowed", "扭亏", "亏损收窄"}:
        return True
    try:
        return float(raw) > 0
    except ValueError:
        match = re.search(r"-?\d+(?:\.\d+)?", raw.replace(",", ""))
        return bool(match and float(match.group(0)) > 0)


def _is_present(value: str) -> bool:
    raw = str(value or "").strip().lower()
    return bool(raw and raw != "missing")


def _improvement_signals(event: dict[str, str]) -> int:
    signals = 0
    signals += int(_positive_number(event.get("revenue_yoy", "")))
    signals += int(_positive_number(event.get("product_revenue_yoy", "")))
    profit_yoy = event.get("profit_yoy", "").lower()
    signals += int(_positive_number(profit_yoy) or profit_yoy in {"turnaround", "loss_narrowed", "亏损收窄", "扭亏"})
    signals += int(_boolish(event.get("business_improved", "")))
    signals += int(_boolish(event.get("loss_narrowed", "")))
    signals += int(_boolish(event.get("turned_profitable", "")))
    signals += int(_boolish(event.get("guidance_raised", "")))
    return signals


def _earnings_objective(defn: S2Definition, earnings_events: list[dict[str, str]]) -> S2Item:
    active = [event for event in earnings_events if _is_confirmed_active(event)]
    if not active:
        adjusted, confidence = _adjust(0.5, 0, missing="业绩事件库为空")
        return S2Item(defn.code, defn.name, None, 0.5, adjusted, confidence, "数据缺失", "事件库无有效财报客观改善记录", "earnings_events.csv", 0, 0, "财报客观改善样本缺失")
    improved = sum(1 for event in active if _improvement_signals(event) >= 2)
    value = improved / len(active)
    raw_score, rating = _rate_score(value, 0.70, 0.40)
    adjusted, confidence = _adjust(raw_score, len(active))
    if len(active) < 3 and value >= 0.70:
        rating = "positive_low_sample"
    basis = f"财报客观改善样本 {improved}/{len(active)}；仅基于同比、利润改善、亏损收窄、现金流/经营改善等客观项，不代表超一致预期；样本不足3个时不得写超预期"
    return S2Item(defn.code, defn.name, value, raw_score, adjusted, confidence, rating, basis, "earnings_events.csv", len(active))


EARNINGS_CONSENSUS_FIELDS = [
    "company_name",
    "symbol",
    "report_period",
    "actual_revenue",
    "consensus_revenue",
    "revenue_beat",
    "actual_adjusted_profit",
    "consensus_adjusted_profit",
    "profit_beat",
    "actual_eps",
    "consensus_eps",
    "eps_beat",
    "consensus_source",
    "source_url",
    "source_date",
    "confidence",
    "note",
]


EARNINGS_CONSENSUS_CORE_ROWS = [
    ("06160.HK", "百济神州"),
    ("01801.HK", "信达生物"),
    ("09926.HK", "康方生物"),
    ("600276.SH", "恒瑞医药"),
    ("603259.SH", "药明康德"),
    ("02269.HK", "药明生物"),
]


def ensure_earnings_consensus(data_dir: Path) -> None:
    path = data_dir / "earnings_consensus.csv"
    if path.exists():
        return
    data_dir.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=EARNINGS_CONSENSUS_FIELDS, lineterminator="\n")
        writer.writeheader()
        for symbol, company in EARNINGS_CONSENSUS_CORE_ROWS:
            row = {field: "missing" for field in EARNINGS_CONSENSUS_FIELDS}
            row.update({
                "company_name": company,
                "symbol": symbol,
                "note": "未接入可靠一致预期来源；不得用同比增长替代 beat/miss",
            })
            writer.writerow(row)


def _has_consensus_source(row: dict[str, str]) -> bool:
    source = row.get("consensus_source") or row.get("source_url") or row.get("consensus_source_url")
    return _is_present(source)


def _beat_value(row: dict[str, str]) -> bool | None:
    fields = ["revenue_beat", "profit_beat", "eps_beat", "beat"]
    values = [row.get(field, "") for field in fields if _is_present(row.get(field, ""))]
    if not values:
        return None
    positive = sum(_boolish(value) for value in values)
    return positive >= max(1, len(values) / 2)


def _earnings_consensus(defn: S2Definition, earnings_events: list[dict[str, str]], consensus_rows: list[dict[str, str]]) -> S2Item:
    active = [
        row for row in consensus_rows
        if _has_consensus_source(row) and _beat_value(row) is not None
    ]
    if not active:
        active = [
            event for event in earnings_events
            if _is_confirmed_active(event)
            and event.get("has_consensus", "").lower() in {"true", "1", "yes", "y", "是"}
            and event.get("beat", "") != ""
            and _has_consensus_source(event)
        ]
    if not active:
        adjusted, confidence = _adjust(0.5, 0, missing="业绩事件库为空")
        return S2Item(defn.code, defn.name, None, 0.5, adjusted, confidence, "数据缺失", "earnings_consensus.csv 无可靠一致预期来源；不得用同比增长冒充超预期", "earnings_consensus.csv", 0, 0, "具备一致预期来源的业绩样本缺失")
    beats = sum(1 for event in active if _beat_value(event))
    value = beats / len(active)
    raw_score, rating = _rate_score(value, 0.60, 0.40)
    adjusted, confidence = _adjust(raw_score, len(active))
    return S2Item(defn.code, defn.name, value, raw_score, adjusted, confidence, rating, f"{beats}/{len(active)} 个具备可靠一致预期来源的业绩事件标记为 beat", "earnings_consensus.csv", len(active))


COMMERCIALIZATION_FIELDS = [
    "company_name",
    "symbol",
    "report_period",
    "total_revenue_yoy",
    "product_revenue_yoy",
    "innovation_drug_revenue_yoy",
    "adjusted_profit_yoy",
    "operating_cash_flow",
    "cash_balance",
    "source_url",
    "source_date",
    "source_type",
    "note",
]


COMMERCIALIZATION_CORE_ROWS = [
    ("06160.HK", "百济神州"),
    ("01801.HK", "信达生物"),
    ("09926.HK", "康方生物"),
    ("600276.SH", "恒瑞医药"),
    ("603259.SH", "药明康德"),
    ("02269.HK", "药明生物"),
]


def ensure_commercialization_metrics(data_dir: Path) -> None:
    path = data_dir / "commercialization_metrics.csv"
    if path.exists():
        return
    data_dir.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=COMMERCIALIZATION_FIELDS, lineterminator="\n")
        writer.writeheader()
        for ticker, company in COMMERCIALIZATION_CORE_ROWS:
            writer.writerow({**{field: "" for field in COMMERCIALIZATION_FIELDS}, "symbol": ticker, "company_name": company, "note": "missing"})


def _commercialization_quality(defn: S2Definition, data_dir: Path) -> S2Item:
    ensure_commercialization_metrics(data_dir)
    rows = load_events(data_dir / "commercialization_metrics.csv")
    metric_fields = [field for field in COMMERCIALIZATION_FIELDS if field not in {"ticker", "company", "company_name", "symbol", "report_period", "source_url", "source_date", "source_type", "note"}]
    usable = [row for row in rows if any(_is_present(row.get(field, "")) for field in metric_fields)]
    if not usable:
        adjusted, confidence = _adjust(0.5, 0, missing="商业化兑现质量数据缺失")
        return S2Item(
            defn.code,
            defn.name,
            None,
            0.5,
            adjusted,
            confidence,
            "数据缺失",
            "commercialization_metrics.csv 已建表，但核心公司商业化指标仍为 missing",
            "commercialization_metrics.csv",
            0,
            0,
            "commercialization_metrics missing",
        )
    passed = 0
    filled_fields = sum(1 for row in rows for field in metric_fields if _is_present(row.get(field, "")))
    total_fields = len(rows) * len(metric_fields)
    completeness = filled_fields / total_fields if total_fields else 0.0
    missing_note = ""
    if len(usable) < len(rows) or completeness < 1.0:
        missing_note = f"商业化核心公司覆盖={len(usable)}/{len(rows)}；关键字段完整度={completeness:.0%}；缺失字段保留missing"
    for row in usable:
        signals = 0
        signals += int(_positive_number(row.get("product_revenue_yoy", "")))
        signals += int(_positive_number(row.get("innovation_drug_revenue_yoy", "")))
        signals += int(_positive_number(row.get("total_revenue_yoy", "")))
        signals += int(_positive_number(row.get("adjusted_profit_yoy", "")))
        signals += int(_positive_number(row.get("operating_cash_flow", "")))
        signals += int(_positive_number(row.get("cash_balance", "")))
        passed += int(signals >= 3)
    value = passed / len(usable)
    raw_score, rating = _rate_score(value, 0.70, 0.40)
    adjusted, confidence = _adjust(raw_score, len(usable), missing=missing_note)
    basis = f"商业化兑现质量通过样本 {passed}/{len(usable)}；只判断收入、利润、现金流和现金余额等商业化质量，不替代S2-03"
    return S2Item(defn.code, defn.name, value, raw_score, adjusted, confidence, rating, basis, "commercialization_metrics.csv", len(usable), 0, missing_note)


def _clinical(defn: S2Definition, clinical_events: list[dict[str, str]], market_data_dir: Path, trade_date: str, audit_path: Path | None = None) -> tuple[S2Item, MarketResult]:
    result = clinical_conversion_rate(clinical_events, market_data_dir, as_of_trade_date=trade_date, audit_path=audit_path)
    missing = _merge_missing(*result.missing)
    raw_score, rating = _rate_score(result.value, 0.60, 0.40)
    if result.value is None and result.pending_count > 0 and result.proxy_sample_count == 0:
        rating = "待验证"
    adjusted, confidence = _adjust(raw_score, result.sample_count, result.replacement_count, missing)
    if result.true_sample_count == 0 and result.proxy_sample_count > 0:
        adjusted = min(adjusted, 0.60)
        confidence = min(confidence, 0.60)
    return S2Item(
        defn.code,
        defn.name,
        result.value,
        raw_score,
        adjusted,
        confidence,
        rating,
        result.basis,
        "clinical_events.csv + 本地行情",
        result.sample_count,
        result.replacement_count,
        missing,
        true_value=result.true_value,
        proxy_value=result.proxy_value,
        true_sample_count=result.true_sample_count,
        proxy_sample_count=result.proxy_sample_count,
        proxy_type=result.proxy_type,
        pending_count=result.pending_count,
        hk_pending_count=result.hk_pending_count,
        price_missing_count=result.price_missing_count,
        raw_mature_event_count=result.raw_mature_event_count,
        deduped_trade_sample_count=result.deduped_trade_sample_count,
        success_count=result.success_count,
        success_rate=result.success_rate,
        clinical_event_statuses=result.clinical_event_statuses,
    ), result


def _leader_tickers(data_dir: Path) -> list[str]:
    path = data_dir / "leader_pool.csv"
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8") as fh:
        return [row["ticker"] for row in csv.DictReader(fh) if row.get("ticker")]


def _leader(defn: S2Definition, catalyst_events: list[dict[str, str]], market_data_dir: Path, data_dir: Path, trade_date: str) -> tuple[S2Item, MarketResult]:
    result = leader_excess_median(catalyst_events, market_data_dir, local_leaders=_leader_tickers(data_dir), as_of_trade_date=trade_date)
    missing = _merge_missing(*result.missing)
    raw_score, rating = _rate_score(result.value, 0.05, 0.0)
    if result.value is None and result.pending_count > 0:
        rating = "待验证"
    adjusted, confidence = _adjust(raw_score, result.sample_count, result.replacement_count, missing)
    return S2Item(
        defn.code,
        defn.name,
        result.value,
        raw_score,
        adjusted,
        confidence,
        rating,
        result.basis,
        "事件库 + leader_pool.csv + 本地行情",
        result.sample_count,
        result.replacement_count,
        missing,
        leader_excess_median_5d=result.leader_excess_median_5d,
        leader_win_rate_5d=result.leader_win_rate_5d,
        leader_excess_median_10d=result.leader_excess_median_10d,
        leader_breadth_20d=result.leader_breadth_20d,
        pending_count=result.pending_count,
    ), result


def _optional_float(value: str) -> float | None:
    try:
        return float(value) if value else None
    except ValueError:
        return None


def _int(value: str) -> int:
    try:
        return int(value or 0)
    except ValueError:
        return 0


def _item_from_row(row: dict[str, str]) -> S2Item:
    return S2Item(
        code=row["code"],
        name=row["name"],
        value=_optional_float(row.get("value", "")),
        raw_score=float(row["raw_score"]),
        adjusted_score=float(row["adjusted_score"]),
        confidence=float(row["confidence"]),
        rating=row["rating"],
        basis=row["basis"],
        source=row["source"],
        sample_count=_int(row.get("sample_count", "")),
        replacement_count=_int(row.get("replacement_count", "")),
        missing=row.get("missing", ""),
        event_db_maturity=row.get("event_db_maturity", ""),
        raw_bd_amount=_optional_float(row.get("raw_bd_amount", "")),
        quality_bd_amount=_optional_float(row.get("quality_bd_amount", "")),
        baseline_bd_amount=_optional_float(row.get("baseline_bd_amount", "")),
        true_value=_optional_float(row.get("true_value", "")),
        proxy_value=_optional_float(row.get("proxy_value", "")),
        true_sample_count=_int(row.get("true_sample_count", "")),
        proxy_sample_count=_int(row.get("proxy_sample_count", "")),
        proxy_type=row.get("proxy_type", ""),
        leader_excess_median_5d=_optional_float(row.get("leader_excess_median_5d", "")),
        leader_win_rate_5d=_optional_float(row.get("leader_win_rate_5d", "")),
        leader_excess_median_10d=_optional_float(row.get("leader_excess_median_10d", "")),
        leader_breadth_20d=_optional_float(row.get("leader_breadth_20d", "")),
        pending_count=_int(row.get("pending_count", "")),
        hk_pending_count=_int(row.get("hk_pending_count", "")),
        price_missing_count=_int(row.get("price_missing_count", "")),
        raw_mature_event_count=_int(row.get("raw_mature_event_count", "")),
        deduped_trade_sample_count=_int(row.get("deduped_trade_sample_count", "")),
        success_count=_int(row.get("success_count", "")),
        success_rate=_optional_float(row.get("success_rate", "")),
        carried_forward_from=row.get("carried_forward_from", ""),
        stale_days=_int(row.get("stale_days", "")),
        is_stale=row.get("is_stale", "").lower() == "true",
        carry_forward_type=row.get("carry_forward_type", ""),
    )


def _load_previous_market_items(output_dir: Path, report_date: str | None) -> tuple[str, str, dict[str, S2Item]]:
    path = output_dir / "s2_item_scores.csv"
    if not report_date or not path.exists():
        return "", "", {}
    with path.open(newline="", encoding="utf-8") as fh:
        rows = [row for row in csv.DictReader(fh) if row.get("date", "") < report_date]
    dates = sorted({row["date"] for row in rows}, reverse=True)
    for date in dates:
        selected = {row["code"]: _item_from_row(row) for row in rows if row["date"] == date and row["code"] in {"S2-04", "S2-05"}}
        if set(selected) == {"S2-04", "S2-05"}:
            trade_date = next(row["s1_trade_date"] for row in rows if row["date"] == date)
            return date, trade_date, selected
    return "", "", {}


def _merge_missing(*values: str) -> str:
    parts: list[str] = []
    for value in values:
        for part in value.split("；"):
            if part and part not in parts:
                parts.append(part)
    return "；".join(parts)


def _trade_day_distance(market_data_dir: Path, start_trade_date: str, end_trade_date: str) -> int:
    path = market_data_dir / "fund_daily.csv"
    if not path.exists():
        return 999
    with path.open(newline="", encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    dates = sorted(
        {
            row.get("trade_date", "")
            for row in rows
            if row.get("ts_code") == "589720.SH" and row.get("trade_date")
        }
    )
    start = start_trade_date.replace("-", "")
    end = end_trade_date.replace("-", "")
    return len([date for date in dates if start < date <= end])


def _carry_forward(current: S2Item, previous: S2Item, previous_report_date: str, stale_days: int) -> S2Item:
    origin = previous.carried_forward_from or previous_report_date
    refresh_trade_sample_counts = current.code == "S2-04"
    base_source = current.basis if refresh_trade_sample_counts and current.value is not None else previous.basis
    base_basis = base_source.split("；本日无新增成熟样本", 1)[0]
    carry_type = "recent_carry_forward" if stale_days <= 2 else "aging_carry_forward"
    cap_text = "" if carry_type == "recent_carry_forward" else "；沿用已超过2个交易日，adjusted_score封顶0.60"
    basis = f"{base_basis}；本日无新增成熟样本，沿用 {origin} 有效观测；待验证 {current.pending_count} 个{cap_text}"
    return replace(
        previous,
        adjusted_score=min(previous.adjusted_score, 0.60) if carry_type == "aging_carry_forward" else previous.adjusted_score,
        basis=basis,
        missing=_merge_missing(previous.missing, current.missing),
        value=current.value if refresh_trade_sample_counts and current.value is not None else previous.value,
        raw_score=current.raw_score if refresh_trade_sample_counts and current.value is not None else previous.raw_score,
        sample_count=current.sample_count if refresh_trade_sample_counts else previous.sample_count,
        replacement_count=current.replacement_count if refresh_trade_sample_counts else previous.replacement_count,
        true_value=current.true_value if refresh_trade_sample_counts else previous.true_value,
        proxy_value=current.proxy_value if refresh_trade_sample_counts else previous.proxy_value,
        true_sample_count=current.true_sample_count if refresh_trade_sample_counts else previous.true_sample_count,
        proxy_sample_count=current.proxy_sample_count if refresh_trade_sample_counts else previous.proxy_sample_count,
        proxy_type=current.proxy_type if refresh_trade_sample_counts else previous.proxy_type,
        pending_count=current.pending_count,
        hk_pending_count=current.hk_pending_count,
        price_missing_count=current.price_missing_count,
        raw_mature_event_count=current.raw_mature_event_count,
        deduped_trade_sample_count=current.deduped_trade_sample_count,
        success_count=current.success_count,
        success_rate=current.success_rate,
        carried_forward_from=origin,
        stale_days=stale_days,
        is_stale=False,
        carry_forward_type=carry_type,
        clinical_event_statuses=current.clinical_event_statuses,
    )


def _expire_carry(current: S2Item, code: str, stale_days: int) -> S2Item:
    return replace(
        current,
        value=None,
        raw_score=0.5,
        adjusted_score=0.5,
        confidence=min(current.confidence, 0.45),
        rating="数据缺失",
        missing=_merge_missing(current.missing, f"最近一次{code}有效观测已超过5个交易日，不再沿用"),
        stale_days=stale_days,
        is_stale=True,
        carry_forward_type="stale",
    )


def _level(score: float) -> str:
    if score >= 0.80:
        return "超预期"
    if score >= 0.60:
        return "符合预期"
    if score >= 0.40:
        return "低于预期"
    return "显著低于预期"


def score_s2(
    trade_date: str,
    data_dir: Path,
    output_dir: Path,
    market_data_dir: Path,
    excel_path: Path,
    s1_total: float | None = None,
    s1_share_change: float | None = None,
    report_date: str | None = None,
) -> S2Score:
    ensure_event_store(data_dir)
    definitions = load_definitions(excel_path)
    bd_events = load_events(data_dir / "bd_events.csv")
    clinical_events = load_events(data_dir / "clinical_events.csv")
    earnings_events = load_events(data_dir / "earnings_events.csv")
    ensure_earnings_consensus(data_dir)
    earnings_consensus = load_events(data_dir / "earnings_consensus.csv")
    regulatory_events = load_events(data_dir / "regulatory_events.csv")
    clinical_validation_events = _market_events_since(clinical_events, trade_date, 90, report_date)
    catalyst_events = _market_events_since(clinical_events + bd_events + regulatory_events, trade_date, 90, report_date)
    audit_path = output_dir / "data_audit" / "market_data_audit.csv"
    clinical_item, clinical_result = _clinical(definitions["S2-04"], clinical_validation_events, market_data_dir, trade_date, audit_path if audit_path.exists() else None)
    leader_item, leader_result = _leader(definitions["S2-05"], catalyst_events, market_data_dir, data_dir, trade_date)

    items = {
        "S2-01": _bd_frequency(definitions["S2-01"], bd_events, trade_date),
        "S2-02": _bd_quality(definitions["S2-02"], bd_events, trade_date),
        "S2-03a": _earnings_objective(definitions["S2-03a"], earnings_events),
        "S2-03b": _earnings_consensus(definitions["S2-03b"], earnings_events, earnings_consensus),
        "S2-04": clinical_item,
        "S2-05": leader_item,
    }
    explanation_items = {
        "S2-06": _commercialization_quality(definitions["S2-06"], data_dir),
    }
    previous_report_date, previous_trade_date, previous_items = _load_previous_market_items(output_dir, report_date)
    if previous_items:
        previous_clinical = clinical_conversion_rate(
            clinical_validation_events,
            market_data_dir,
            as_of_trade_date=previous_trade_date,
            audit_path=audit_path if audit_path.exists() else None,
        )
        previous_leader = leader_excess_median(catalyst_events, market_data_dir, local_leaders=_leader_tickers(data_dir), as_of_trade_date=previous_trade_date)
        elapsed_trade_days = _trade_day_distance(market_data_dir, previous_trade_date, trade_date)
        previous_clinical_item = previous_items["S2-04"]
        clinical_stale_days = previous_clinical_item.stale_days + elapsed_trade_days
        if (
            not set(clinical_result.mature_event_keys) - set(previous_clinical.mature_event_keys)
            and previous_clinical_item.true_sample_count > 0
            and clinical_stale_days <= 5
        ):
            items["S2-04"] = _carry_forward(items["S2-04"], previous_clinical_item, previous_report_date, clinical_stale_days)
        elif not set(clinical_result.mature_event_keys) - set(previous_clinical.mature_event_keys) and clinical_stale_days > 5:
            items["S2-04"] = _expire_carry(items["S2-04"], "S2-04", clinical_stale_days)
        previous_leader_item = previous_items["S2-05"]
        leader_stale_days = previous_leader_item.stale_days + elapsed_trade_days
        if (
            not set(leader_result.mature_event_keys) - set(previous_leader.mature_event_keys)
            and previous_leader_item.sample_count > 0
            and previous_leader_item.replacement_count == 0
            and leader_stale_days <= 5
        ):
            items["S2-05"] = _carry_forward(items["S2-05"], previous_leader_item, previous_report_date, leader_stale_days)
        elif not set(leader_result.mature_event_keys) - set(previous_leader.mature_event_keys) and leader_stale_days > 5:
            items["S2-05"] = _expire_carry(items["S2-05"], "S2-05", leader_stale_days)
    total_weight = sum(definitions[code].weight for code in items)
    raw_score = sum(items[code].raw_score * definitions[code].weight for code in items) / total_weight
    adjusted_score = sum(items[code].adjusted_score * definitions[code].weight for code in items) / total_weight
    missing = [item.missing for item in items.values() if item.missing]
    available_weight = sum(
        definitions[code].weight for code, item in items.items()
        if item.value is not None and item.rating != "数据缺失"
    )
    return S2Score(
        trade_date=trade_date,
        raw_score=raw_score,
        adjusted_score=adjusted_score,
        items=items,
        missing_data=_merge_missing(*missing),
        level=_level(adjusted_score),
        available_weight=available_weight,
        missing_indicator_count=sum(item.rating == "数据缺失" for item in items.values()),
        pending_indicator_count=sum(item.rating == "待验证" for item in items.values()),
        proxy_indicator_count=sum(item.proxy_sample_count > 0 for item in items.values()),
        stale_indicator_count=sum(item.is_stale for item in items.values()),
        explanation_items=explanation_items,
    )
